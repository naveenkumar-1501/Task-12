"""
Excel_function.py
"""

from openpyxl import load_workbook

class Naveen_Excel_Function:
    def __init__(self, excel_file_name, sheet_name):
       # Initialize with the file name and sheet name
       self.file = excel_file_name
       self.sheet = sheet_name

    def row_count(self):
        # Load the workbook and get the sheet by name
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        # Return the maximum number of rows in the sheet
        return sheet.max_row

    def column_count(self):
        # Load the workbook and get the sheet by name
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        # Return the maximum number of columns in the sheet
        return sheet.max_column

    def read_data(self, row_number, column_number):
        # Load the workbook and get the sheet by name
        workbook = load_workbook(self.file)
        sheet = workbook [self.sheet]
        # Read and return the data from the specified cell
        return sheet.cell(row=row_number, column=column_number).value

    def write_data(self, row_number, column_number, data):
        # Load the workbook and get the sheet by name
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        # Write data to the specified cell
        sheet.cell(row=row_number, column=column_number).value = data
        # Save the workbook after writing data
        workbook.save(self.file)

